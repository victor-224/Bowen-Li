"""Flask API for industrial digital twin auto-ingestion pipeline."""

from __future__ import annotations

from pathlib import Path
from urllib import error as urlerror
from urllib import parse as urlparse
from urllib import request as urlrequest
import json
import logging
import os
import socket
import threading
import time
from dataclasses import dataclass
from typing import Any, Dict, List, MutableMapping, Optional

import openpyxl
from flask import Flask, jsonify, request
from flask_cors import CORS
from backend.asset_contract import AssetContractViolation
from backend.file_classifier import classify_files
from backend.pdf_loader import first_page_to_layout_png
from backend.relations import build_relations
from backend.walls import parse_walls_and_rooms
from backend.layout_graph import build_layout_graph
from backend.multiplant_registry import list_plants, register_snapshot
from backend.observability import audit_event, finish_trace, get_observability, observe_operation, start_trace
from backend.pid_integration import build_pid_links
from backend.ai_service import call_copilot_model
from backend.config import (
    AI_RUNTIME_CONFIG_FILE,
    get_ai_model_defaults,
    get_lm_studio_chat_url,
    lm_studio_url_from_env,
)
from backend.core.policy_engine import resolve_policy
from backend.core.spatial_truth_ledger import summarize_truth_status
from backend.core.spatial_payload import sanitize_spatial_payload
from backend.models.vision.vision_schema import normalize_vision_output
from backend.models.vision.vl_interface import run_vision_model
from backend.runtime_state import RuntimeState
from backend.topology_optimizer import optimize_topology
from backend.plan_upload_validate import validate_layout_raster
from backend.layout_inspector import inspect_runtime_layout

SHEET_NAME = "Equipment_list"
# Real annex: column B = TAG, C = SERVICE, D = POSITION (or TEMA), E/F/G = dimensions (mm)
DATA_START_ROW = 9
COL_TAG = 2
COL_SERVICE = 3
COL_POSITION = 4
COL_DIAMETER = 5
COL_LENGTH = 6
COL_HEIGHT = 7

_EXCEL_REL = Path("data") / "Copy of Annexe 2_Equipment_liste_et_taille.xlsx"
_RUNTIME_DIR_REL = Path("data") / "runtime"
_RUNTIME_PLAN_REL = _RUNTIME_DIR_REL / "plan.png"
_RUNTIME_EXCEL_REL = _RUNTIME_DIR_REL / "equipment.xlsx"
MAX_UPLOAD_BYTES = 80 * 1024 * 1024  # 80MB soft limit for demo stability
PIPELINE_TIMEOUT_SECONDS = 120
# Optional VLM enrichment (set ENABLE_VISION=true in environment to enable).
ENABLE_VISION = False
_VISION_INTEGRATION_LOG = logging.getLogger("industrial_digital_twin.vision.integration")
_VISION_INTEGRATION_LOGGED = False
_VISION_MODEL = "qwen2.5-vl-7b-instruct"
_VISION_PROMPT = (
    "Analyze industrial layout. Detect equipment, structures, labels and relations."
)
_VISION_TIMEOUT_S = 20

_RUNTIME_STATE = RuntimeState()
_UPLOAD_LOG = logging.getLogger("industrial_digital_twin.upload")

# Reduce OpenCV console noise (libpng may still print at ERROR level on corrupt reads).
try:
    import cv2.utils.logging as _cv2_log

    _cv2_log.setLogLevel(_cv2_log.LOG_LEVEL_SILENT)
except Exception:  # noqa: BLE001
    pass

TASK_QUEUED = "queued"
TASK_VALIDATING = "validating"
TASK_PROCESSING_OCR = "processing_ocr"
TASK_PARSING_LAYOUT = "parsing_layout"
TASK_BUILDING_GRAPH = "building_graph"
TASK_RENDERING_SCENE = "rendering_scene"
TASK_FINALIZING = "finalizing"
TASK_DONE = "done"
TASK_FAILED = "failed"
TASK_CANCELLED = "cancelled"


@dataclass
class PipelineContext:
    task_id: str
    signature: str
    stage: str = TASK_QUEUED
    cancelled: threading.Event | None = None


_TASK_CONTEXTS: Dict[str, PipelineContext] = {}
_TASK_CTX_LOCK = threading.Lock()
_TASK_CTX_TTL_S = 8 * 60
_TASK_CTX_MAX = 20
_CACHE_STATS: Dict[str, int] = {"hit": 0, "miss": 0}


def _prune_task_contexts() -> None:
    now = time.time()
    with _TASK_CTX_LOCK:
        stale = [k for k, v in _TASK_CONTEXTS.items() if (now - v.cancelled._when if getattr(v.cancelled, "_when", None) else now - 0) > _TASK_CTX_TTL_S]  # type: ignore[attr-defined]
        for k in stale:
            _TASK_CONTEXTS.pop(k, None)
        if len(_TASK_CONTEXTS) > _TASK_CTX_MAX:
            keys = list(_TASK_CONTEXTS.keys())
            for k in keys[: len(_TASK_CONTEXTS) - _TASK_CTX_MAX]:
                _TASK_CONTEXTS.pop(k, None)


def _register_task_context(task_id: str, signature: str) -> PipelineContext:
    ev = threading.Event()
    ctx = PipelineContext(task_id=task_id, signature=signature, cancelled=ev)
    with _TASK_CTX_LOCK:
        _TASK_CONTEXTS[task_id] = ctx
    return ctx


def _get_task_context(task_id: str) -> Optional[PipelineContext]:
    with _TASK_CTX_LOCK:
        return _TASK_CONTEXTS.get(task_id)


def _transition(task_id: str, status: str, progress: int, message: str) -> None:
    current = _RUNTIME_STATE.get_task(task_id) or {}
    current_stage = str(current.get("stage") or status)
    next_stage = current_stage
    if status in {
        TASK_QUEUED,
        TASK_VALIDATING,
        TASK_PROCESSING_OCR,
        TASK_PARSING_LAYOUT,
        TASK_BUILDING_GRAPH,
        TASK_RENDERING_SCENE,
        TASK_FINALIZING,
    }:
        next_stage = status
    _RUNTIME_STATE.update_task(task_id, status=status, stage=next_stage, progress=progress, message=message)
    ctx = _get_task_context(task_id)
    if ctx is not None:
        ctx.stage = next_stage
    audit_event(runtime_dir_path(), "task_transition", {"task_id": task_id, "status": status, "message": message})


def _cancelled(ctx: PipelineContext | None) -> bool:
    if ctx is None or ctx.cancelled is None:
        return False
    return ctx.cancelled.is_set()


def _pipeline_error(code: str, message: str, stage: str) -> Dict[str, Any]:
    return {"code": code, "message": message, "stage": stage}


def _error_response_struct(code: str, message: str, stage: str, status_code: int = 500) -> Any:
    return jsonify({"success": False, "error": _pipeline_error(code, message, stage)}), status_code


def _asset_violation_response(e: AssetContractViolation, status_code: int = 400) -> Any:
    payload = e.to_error_payload()
    return jsonify({"success": False, "error": payload}), status_code


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def runtime_dir_path() -> Path:
    return _repo_root() / _RUNTIME_DIR_REL


def runtime_plan_path() -> Path:
    return _repo_root() / _RUNTIME_PLAN_REL


def runtime_excel_path() -> Path:
    return _repo_root() / _RUNTIME_EXCEL_REL


def data_dir_path() -> Path:
    return _repo_root() / "data"


def _path_like_to_json(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return [str(x) for x in value]
    return None


def _error_response(message: str, status_code: int = 500) -> Any:
    return _error_response_struct("PIPELINE_ERROR", message, "unknown", status_code)


def _pipeline_signature() -> str:
    classified = classify_files(data_dir_path())
    paths = {
        "runtime_plan": runtime_plan_path() if runtime_plan_path().is_file() else None,
        "runtime_excel": runtime_excel_path() if runtime_excel_path().is_file() else None,
        "layout": Path(str(classified["layout"])) if classified.get("layout") else None,
        "excel": Path(str(classified["excel"])) if classified.get("excel") else None,
        "reference": Path(str(classified["reference"])) if classified.get("reference") else None,
        "gad": Path(str(classified["gad"])) if classified.get("gad") else None,
        "structure": Path(str(classified["structure"])) if classified.get("structure") else None,
    }
    return _RUNTIME_STATE.build_signature(paths)


def _positions_cache_path() -> Path:
    return runtime_dir_path() / "positions_cache.json"


def _invalidate_spatial_runtime_cache(*, clear_positions: bool = False) -> Dict[str, Any]:
    """Invalidate runtime caches that can preserve stale spatial state."""
    cleared_payloads = _RUNTIME_STATE.clear_pipeline_cache()
    cleared_positions = False
    if clear_positions:
        p = _positions_cache_path()
        if p.is_file():
            try:
                p.unlink()
                cleared_positions = True
            except OSError:
                cleared_positions = False
    return {"payload_cache_cleared": int(cleared_payloads), "positions_cache_cleared": bool(cleared_positions)}


def _layout_upload_validation_message(reason: str) -> str:
    r = (reason or "").strip()
    if r == "invalid_png_signature":
        return (
            "Uploaded layout is not a valid PNG file (wrong file header). "
            "Export the drawing as PNG or upload PDF instead."
        )
    if r == "invalid_jpeg_signature":
        return "Uploaded layout is not a valid JPEG file. Re-export the image or use PDF/PNG."
    if r == "opencv_decode_failed":
        return (
            "Layout image could not be decoded (file may be truncated, renamed, or not an image). "
            "Re-upload a real PNG/JPEG or use PDF."
        )
    if r == "file_is_pdf_not_image":
        return "Layout file was saved as image but content looks like PDF. Upload it as plan_file with .pdf extension."
    return f"Layout validation failed ({r or 'unknown'})."


def _get_or_build_pipeline_sync(equipment: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    signature = _pipeline_signature()
    cached = _RUNTIME_STATE.get_cached_payload(signature)
    if cached is not None:
        _CACHE_STATS["hit"] += 1
        return sanitize_spatial_payload(cached, route="cache_hit")
    _CACHE_STATS["miss"] += 1
    payload = build_pipeline_output(equipment)
    payload = sanitize_spatial_payload(payload, route="cache_store")
    _RUNTIME_STATE.set_cached_payload(signature=signature, payload=payload)
    return payload


def _classify_pipeline_error(message: str) -> str:
    m = (message or "").lower()
    if "sheet" in m and "equipment_list" in m:
        return "Excel format invalid: sheet 'Equipment_list' is required."
    if "excel not found" in m:
        return "No Excel detected. Upload a valid .xlsx equipment list."
    if "[scene_error] plan.png missing" in m or "plan image not found" in m:
        return "Scene generation failed due to missing plan.png asset."
    if "[scene_error] plan.png is corrupted" in m or "cannot read plan image" in m:
        return "Scene generation failed due to corrupted plan.png asset."
    if "no positions detected" in m:
        return "OCR failed to detect equipment tags from layout."
    if "upload too large" in m:
        return "Uploaded file is too large for demo mode."
    return message


def _classify_pipeline_error_code(message: str, default: str = "PIPELINE_ERROR") -> tuple[str, str]:
    """Map raw exception message to (error_code, stage)."""
    m = (message or "").lower()
    if "[scene_error] plan.png missing" in m or "plan image not found" in m:
        return "MISSING_PLAN_IMAGE", TASK_RENDERING_SCENE
    if "[scene_error] plan.png is corrupted" in m or "cannot read plan image" in m:
        return "CORRUPTED_PLAN_IMAGE", TASK_RENDERING_SCENE
    if "sheet" in m and "equipment_list" in m:
        return "INVALID_EXCEL", TASK_VALIDATING
    if "excel not found" in m or "file is not a zip file" in m:
        return "INVALID_EXCEL", TASK_VALIDATING
    if "no positions detected" in m:
        return "OCR_FAILED", TASK_PROCESSING_OCR
    return default, TASK_VALIDATING


def _excel_path() -> Path:
    runtime_excel = runtime_excel_path()
    if runtime_excel.is_file():
        return runtime_excel
    classified = classify_files(data_dir_path())
    xlsx = classified.get("excel")
    if xlsx:
        return Path(str(xlsx))
    return _repo_root() / _EXCEL_REL


def plan_image_path() -> Path:
    runtime_plan = runtime_plan_path()
    if runtime_plan.is_file():
        return runtime_plan
    classified = classify_files(data_dir_path())
    layout_src = classified.get("layout")
    if layout_src:
        p = Path(str(layout_src))
        if p.suffix.lower() == ".pdf":
            runtime_dir_path().mkdir(parents=True, exist_ok=True)
            return first_page_to_layout_png(p, runtime_plan)
        return p
    return _repo_root() / Path("data") / "plan_hd.png"


def _vision_enabled() -> bool:
    """Read ENABLE_VISION: env overrides module constant when set."""
    raw = os.environ.get("ENABLE_VISION")
    if raw is None or raw == "":
        return ENABLE_VISION
    return str(raw).strip().lower() in {"1", "true", "yes"}


def _maybe_attach_vision(payload: Dict[str, Any]) -> None:
    """After successful pipeline dict, optionally add payload['vision'] (fail-safe)."""
    global _VISION_INTEGRATION_LOGGED
    if not _vision_enabled():
        if not _VISION_INTEGRATION_LOGGED:
            _VISION_INTEGRATION_LOG.info("vision integration disabled (ENABLE_VISION is off or unset)")
            _VISION_INTEGRATION_LOGGED = True
        return
    if not _VISION_INTEGRATION_LOGGED:
        _VISION_INTEGRATION_LOG.info("vision integration enabled (ENABLE_VISION is on)")
        _VISION_INTEGRATION_LOGGED = True
    try:
        raw = run_vision_model(
            image_path=str(plan_image_path()),
            prompt=_VISION_PROMPT,
            model=_VISION_MODEL,
            timeout=_VISION_TIMEOUT_S,
        )
        norm = normalize_vision_output(raw, _VISION_MODEL)
        payload["vision"] = norm
    except Exception as e:  # noqa: BLE001
        _VISION_INTEGRATION_LOG.warning(
            "vision enrichment failed, pipeline continued: %r", e
        )
        payload["vision"] = {
            "objects": [],
            "relations": [],
            "metadata": {
                "model": "disabled_or_failed",
                "confidence": 0.0,
            },
        }


def _normalize_tag(value: object) -> str:
    """Collapse whitespace so e.g. 'P202 A' matches pickpoint tag 'P202A'."""
    if value is None:
        return ""
    return "".join(str(value).split())


def _json_scalar(value: object) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    return str(value)


def load_equipment_from_excel(path: Optional[Path] = None) -> Dict[str, Dict[str, Any]]:
    """Load Equipment_list: fixed columns, data from DATA_START_ROW; keys are normalized tags."""
    p = path if path is not None else _excel_path()
    if not p.is_file():
        raise FileNotFoundError(f"Excel not found: {p}")

    out: Dict[str, Dict[str, Any]] = {}
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet {SHEET_NAME!r} not found; have {wb.sheetnames!r}")
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(
            min_row=DATA_START_ROW,
            min_col=COL_TAG,
            max_col=COL_HEIGHT,
            values_only=True,
        ):
            tag_raw = row[0]
            tag = _normalize_tag(tag_raw)
            if not tag:
                continue
            entry: MutableMapping[str, Any] = {
                "service": _json_scalar(row[1]),
                "position": _json_scalar(row[2]),
                "diameter": _json_scalar(row[3]),
                "length": _json_scalar(row[4]),
                "height": _json_scalar(row[5]),
            }
            out[tag] = dict(entry)
    finally:
        wb.close()
    return out


def equipment_dict_to_list(equipment: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Stable list form: one object per row with explicit tag field."""
    return [{"tag": tag, **row} for tag, row in equipment.items()]


def build_scene(equipment: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    """Unified scene document via core engines (layout + geometry)."""
    from backend.engines.scene import build_scene_document

    if equipment is None:
        equipment = load_equipment_from_excel()
    return build_scene_document(equipment, plan_path=plan_image_path())


def build_pipeline_output(equipment: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    """
    Unified digital-twin pipeline output.

    Flow:
      OCR位置识别 + Excel属性匹配 + 墙体解析 + 关系计算 -> unified payload
    """
    trace = start_trace("build_pipeline")
    status = "ok"
    try:
        if equipment is None:
            equipment = load_equipment_from_excel()
        scene_doc = build_scene(equipment)
        relations = build_relations(scene_doc)
        layout_graph = build_layout_graph(scene_doc, scene_doc.get("walls", []), relations, equipment)
        walls_doc = {
            "walls": scene_doc.get("walls", []),
            "rooms": scene_doc.get("rooms", []),
            "center": scene_doc.get("center", [0.0, 0.0]),
        }
        data_dir = data_dir_path()
        runtime_dir = runtime_dir_path()
        pid_links = build_pid_links(layout_graph, data_dir)
        topology = optimize_topology(layout_graph)
        source_files = classify_files(data_dir)
        plant_version = register_snapshot(
            runtime_dir,
            plant_id="default-plant",
            payload={
                "scene": scene_doc.get("equipment", []),
                "layout_graph": layout_graph,
            },
            source_files=source_files,
        )
        audit_event(
            runtime_dir,
            "pipeline_built",
            {
                "scene_count": len(scene_doc.get("equipment", [])),
                "zone_count": len(layout_graph.get("zones", [])),
                "version_id": plant_version.get("version_id"),
            },
        )
        payload: Dict[str, Any] = {
            "scene": scene_doc.get("equipment", []),
            "relations": relations,
            "walls": walls_doc,
            "layout_graph": layout_graph,
            "layout_inspector": inspect_runtime_layout(runtime_plan_path()),
            "spatial_contract": scene_doc.get("meta", {}).get("spatial_contract", {}),
            "policy": resolve_policy(
                contract=scene_doc.get("meta", {}).get("spatial_contract", {}),
                ai_status={"available": bool(scene_doc.get("meta", {}).get("execution_policy", {}).get("ai_usage_level") != "disabled")},
                runtime_context={
                    "scene_mode_hint": scene_doc.get("meta", {}).get("execution_policy", {}).get("scene_mode"),
                    "input_state": scene_doc.get("meta", {}).get("input_state"),
                },
            ),
            "spatial_truth_status": summarize_truth_status(runtime_dir_path()),
            "phase_c": {
                "pid_links": pid_links,
                "topology_optimization": topology,
                "multiplant_version": plant_version,
            },
        }
        _maybe_attach_vision(payload)
        return sanitize_spatial_payload(payload, route="build_pipeline_output")
    except Exception:
        status = "error"
        raise
    finally:
        trace_result = finish_trace(trace, status=status)
        observe_operation(runtime_dir_path(), trace_result)


def _build_pipeline_dag(ctx: PipelineContext) -> Dict[str, Any]:
    task_id = ctx.task_id
    _transition(task_id, TASK_VALIDATING, 8, "Validating inputs")
    if _cancelled(ctx):
        raise RuntimeError("Task cancelled")
    equipment = load_equipment_from_excel()
    _transition(task_id, TASK_PROCESSING_OCR, 28, "Running OCR stage")
    if _cancelled(ctx):
        raise RuntimeError("Task cancelled")
    scene_doc = build_scene(equipment)
    _transition(task_id, TASK_PARSING_LAYOUT, 45, "Parsing layout/walls")
    if _cancelled(ctx):
        raise RuntimeError("Task cancelled")
    relations = build_relations(scene_doc)
    _transition(task_id, TASK_BUILDING_GRAPH, 65, "Building graph")
    if _cancelled(ctx):
        raise RuntimeError("Task cancelled")
    layout_graph = build_layout_graph(scene_doc, scene_doc.get("walls", []), relations, equipment)
    _transition(task_id, TASK_RENDERING_SCENE, 82, "Rendering scene payload")
    if _cancelled(ctx):
        raise RuntimeError("Task cancelled")
    walls_doc = {
        "walls": scene_doc.get("walls", []),
        "rooms": scene_doc.get("rooms", []),
        "center": scene_doc.get("center", [0.0, 0.0]),
    }
    data_dir = data_dir_path()
    runtime_dir = runtime_dir_path()
    pid_links = build_pid_links(layout_graph, data_dir)
    topology = optimize_topology(layout_graph)
    source_files = classify_files(data_dir)
    plant_version = register_snapshot(
        runtime_dir,
        plant_id="default-plant",
        payload={"scene": scene_doc.get("equipment", []), "layout_graph": layout_graph},
        source_files=source_files,
    )
    _transition(task_id, TASK_FINALIZING, 95, "Finalizing")
    payload = {
        "scene": scene_doc.get("equipment", []),
        "relations": relations,
        "walls": walls_doc,
        "layout_graph": layout_graph,
        "layout_inspector": inspect_runtime_layout(runtime_plan_path()),
        "spatial_contract": scene_doc.get("meta", {}).get("spatial_contract", {}),
        "policy": resolve_policy(
            contract=scene_doc.get("meta", {}).get("spatial_contract", {}),
            ai_status={"available": bool(scene_doc.get("meta", {}).get("execution_policy", {}).get("ai_usage_level") != "disabled")},
            runtime_context={
                "scene_mode_hint": scene_doc.get("meta", {}).get("execution_policy", {}).get("scene_mode"),
                "input_state": scene_doc.get("meta", {}).get("input_state"),
            },
        ),
        "spatial_truth_status": summarize_truth_status(runtime_dir_path()),
        "phase_c": {
            "pid_links": pid_links,
            "topology_optimization": topology,
            "multiplant_version": plant_version,
        },
    }
    _maybe_attach_vision(payload)
    payload = sanitize_spatial_payload(payload, route="build_pipeline_dag")
    _RUNTIME_STATE.set_cached_payload(signature=ctx.signature, payload=payload)
    return payload


app = Flask(__name__)
CORS(
    app,
    resources={r"/api/*": {"origins": ["http://localhost:3000", "http://127.0.0.1:3000"]}},
    supports_credentials=True,
)

_COPILOT_LOG = logging.getLogger("industrial_digital_twin.copilot")
_AI_CFG_LOG = logging.getLogger("industrial_digital_twin.ai_config")
_COPILOT_SYSTEM = (
    "You are a concise technical assistant for industrial plant layout, equipment lists, and P&ID-style "
    "drawings. Answer in plain language. If you lack plant-specific data, say so and give general guidance."
)
_COPILOT_CONTEXT_BUDGET = 12000
_COPILOT_MSG_BUDGET = 4000


def _lm_studio_models_url(chat_url: str) -> str:
    u = (chat_url or "").strip().rstrip("/")
    if u.endswith("/v1/chat/completions"):
        return u[: -len("/chat/completions")] + "/models"
    if "/v1/" in u:
        return u.split("/v1/", 1)[0] + "/v1/models"
    return u + "/v1/models"


def _fetch_lm_models_list(chat_url: str, timeout: int = 8) -> List[str]:
    models_url = _lm_studio_models_url(chat_url)
    try:
        req = urlrequest.Request(models_url, method="GET", headers={"Accept": "application/json"})
        with urlrequest.urlopen(req, timeout=timeout) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
    except (urlerror.URLError, urlerror.HTTPError, TimeoutError, socket.timeout, json.JSONDecodeError, OSError):
        return []
    if not isinstance(raw, dict):
        return []
    data = raw.get("data")
    if not isinstance(data, list):
        return []
    out: List[str] = []
    for item in data:
        if isinstance(item, dict) and isinstance(item.get("id"), str):
            out.append(str(item["id"]))
    return out


def _load_ai_runtime_file() -> Dict[str, Any]:
    p = AI_RUNTIME_CONFIG_FILE
    if not p.is_file():
        return {}
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return d if isinstance(d, dict) else {}
    except (OSError, json.JSONDecodeError, TypeError):
        return {}


def _save_ai_runtime_file(payload: Dict[str, Any]) -> None:
    AI_RUNTIME_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    AI_RUNTIME_CONFIG_FILE.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def _ai_endpoint_display(chat_url: str) -> str:
    try:
        p = urlparse.urlparse(chat_url)
        host = p.hostname or ""
        port = p.port
        if port:
            return f"{host}:{port}"
        return host or chat_url[:80]
    except Exception:  # noqa: BLE001
        return chat_url[:80]


@app.get("/api/ai/config")
def get_ai_config() -> Any:
    """Return effective AI URL (from env + optional runtime override) and model defaults."""
    rt = _load_ai_runtime_file()
    url = get_lm_studio_chat_url()
    models = get_ai_model_defaults()
    return jsonify(
        {
            "success": True,
            "lm_studio_url": url,
            "lm_studio_url_env": lm_studio_url_from_env(),
            "runtime_override": bool((rt.get("lm_studio_url") or "").strip()),
            "models": models,
            "endpoint_display": _ai_endpoint_display(url),
        }
    )


@app.post("/api/ai/config")
def post_ai_config() -> Any:
    """Persist optional runtime override for LM Studio URL and per-role models."""
    data = request.get_json(silent=True) or {}
    url = (data.get("lm_studio_url") or "").strip()
    if url and not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"success": False, "error": "URL must start with http:// or https://"}), 400
    current = _load_ai_runtime_file()
    if url:
        current["lm_studio_url"] = url
    else:
        current.pop("lm_studio_url", None)
    for key in ("vision", "copilot", "reasoning"):
        fk = f"model_{key}"
        v = data.get(fk) if fk in data else data.get(key)
        if isinstance(v, str):
            if v.strip():
                current[fk] = v.strip()
            elif fk in data or key in data:
                current.pop(fk, None)
        elif v is None and (fk in data or key in data):
            current.pop(fk, None)
    _save_ai_runtime_file(current)
    _AI_CFG_LOG.info("ai runtime config updated keys=%s", list(current.keys()))
    return jsonify({"success": True, "lm_studio_url": get_lm_studio_chat_url(), "models": get_ai_model_defaults()})


@app.post("/api/ai/test-connection")
def post_ai_test_connection() -> Any:
    """Probe LM Studio: optional ``lm_studio_url`` in JSON body; else use configured URL."""
    data = request.get_json(silent=True) or {}
    url = (data.get("lm_studio_url") or "").strip() or get_lm_studio_chat_url()
    if not url.startswith("http://") and not url.startswith("https://"):
        return (
            jsonify(
                {
                    "status": "error",
                    "error": "INVALID_URL",
                    "models": [],
                    "success": False,
                    "fallback": True,
                }
            ),
            200,
        )
    models = _fetch_lm_models_list(url, timeout=8)
    if models:
        return jsonify({"status": "connected", "models": models, "success": True})
    probe = call_copilot_model(
        [
            {"role": "system", "content": "Reply with OK only."},
            {"role": "user", "content": "ping"},
        ],
        temperature=0.0,
        timeout=6,
        max_tokens=4,
        base_url=url,
    )
    if probe.get("success"):
        return jsonify(
            {
                "status": "connected",
                "models": [str(probe.get("model") or get_ai_model_defaults().get("copilot", ""))],
                "success": True,
            }
        )
    err = str(probe.get("error") or "AI_OFFLINE")
    return jsonify(
        {
            "status": "offline",
            "models": [],
            "success": False,
            "error": err if "OFFLINE" in err.upper() or "FAILED" in err.upper() else "AI_OFFLINE",
            "fallback": True,
        }
    )


@app.get("/api/ai/status")
def get_ai_status() -> Any:
    """Lightweight status for UI: endpoint display + copilot model + quick reachability."""
    url = get_lm_studio_chat_url()
    md = get_ai_model_defaults()
    models = _fetch_lm_models_list(url, timeout=4)
    connected = bool(models)
    if not connected:
        p = call_copilot_model(
            [{"role": "user", "content": "ping"}],
            temperature=0.0,
            timeout=4,
            max_tokens=2,
        )
        connected = bool(p.get("success"))
    return jsonify(
        {
            "success": True,
            "endpoint": _ai_endpoint_display(url),
            "lm_studio_url": url,
            "copilot_model": md.get("copilot", ""),
            "vision_model": md.get("vision", ""),
            "reasoning_model": md.get("reasoning", ""),
            "status": "connected" if connected else "offline",
        }
    )


@app.post("/api/agent/execute")
def post_agent_execute() -> Any:
    """Whitelisted tool orchestration (internal plan + execute). Does not replace async upload."""
    from backend.agent.tool_orchestrator import execute_intent

    body = request.get_json(silent=True) or {}
    ctx = body.get("context") if isinstance(body.get("context"), dict) else body
    if not isinstance(ctx, dict):
        ctx = {}
    return jsonify(execute_intent(ctx, runtime_state=_RUNTIME_STATE))


def _serialize_copilot_context(obj: Any) -> str:
    """Render optional client-provided project snapshot; trim to budget."""
    if obj is None or obj == {} or obj == []:
        return ""
    try:
        if isinstance(obj, str):
            s = obj.strip()
        else:
            s = json.dumps(obj, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        s = str(obj)[:_COPILOT_CONTEXT_BUDGET]
    if len(s) > _COPILOT_CONTEXT_BUDGET:
        return s[: _COPILOT_CONTEXT_BUDGET - 3] + "..."
    return s


@app.post("/api/copilot")
def post_copilot() -> Any:
    """Local LLM chat for the UI; never 500 for LM Studio offline (fail-soft JSON).

    Accepts ``project_context`` (dict/list/str) with equipment, graph, scene, vision, errors.
    Merged into the user message so the model prefers project facts over generic advice.
    """
    try:
        data = request.get_json(silent=True) or {}
        user_msg = (data.get("message") or "").strip()
        if not user_msg:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "EMPTY_MESSAGE",
                        "content": "",
                    }
                ),
                200,
            )
    except Exception as e:  # noqa: BLE001
        _COPILOT_LOG.warning("copilot bad request: %r", e)
        return (
            jsonify(
                {
                    "success": False,
                    "error": "INVALID_JSON",
                    "content": "",
                }
            ),
            200,
        )
    try:
        ctx_raw = data.get("project_context")
        ctx_str = _serialize_copilot_context(ctx_raw)
        if ctx_str:
            user_content = (
                "When answering, prefer concrete facts from PROJECT DATA (JSON) below. "
                "If data is missing, say so briefly and use general industry guidance only where needed.\n\n"
                "--- PROJECT DATA (JSON) ---\n"
                f"{ctx_str}\n"
                "--- END ---\n\n"
                f"User question:\n{user_msg[:_COPILOT_MSG_BUDGET]}"
            )
        else:
            user_content = user_msg[:8000]
        messages: List[Dict[str, Any]] = [
            {"role": "system", "content": _COPILOT_SYSTEM},
            {"role": "user", "content": user_content},
        ]
        out = call_copilot_model(messages, temperature=0.3, timeout=45, max_tokens=1024)
        if not out.get("success"):
            code = out.get("error", "COPILOT_ERROR")
            _COPILOT_LOG.warning("copilot LM: %s", code)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": code,
                        "content": "",
                        "fallback": bool(out.get("fallback")),
                    }
                ),
                200,
            )
        content = str(out.get("content") or "").strip()
        return jsonify({"success": True, "content": content, "error": None})
    except Exception as e:  # noqa: BLE001
        _COPILOT_LOG.warning("copilot exception: %r", e)
        return (
            jsonify(
                {
                    "success": False,
                    "error": "COPILOT_EXCEPTION",
                    "content": "",
                }
            ),
            200,
        )


@app.get("/api/equipment")
def get_equipment() -> Any:
    try:
        data = load_equipment_from_excel()
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except Exception as e:  # noqa: BLE001
        code, stage = _classify_pipeline_error_code(str(e))
        return _error_response_struct(code, _classify_pipeline_error(str(e)), stage, 400)
    return jsonify(data)


@app.get("/api/scene")
def get_scene() -> Any:
    try:
        equipment = load_equipment_from_excel()
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except Exception as e:  # noqa: BLE001 - structured error output
        code, stage = _classify_pipeline_error_code(str(e))
        return _error_response_struct(code, _classify_pipeline_error(str(e)), stage, 400)
    try:
        pipeline = _get_or_build_pipeline_sync(equipment)
        return jsonify({"equipment": pipeline.get("scene", []), "walls": pipeline.get("walls", {}).get("walls", [])})
    except AssetContractViolation as e:
        return _asset_violation_response(e, status_code=400)
    except FileNotFoundError as e:
        code, stage = _classify_pipeline_error_code(str(e))
        return _error_response_struct(code, _classify_pipeline_error(str(e)), stage, 400)
    except Exception as e:  # noqa: BLE001 - structured error output
        code, stage = _classify_pipeline_error_code(str(e))
        return _error_response_struct(code, _classify_pipeline_error(str(e)), stage, 500)


@app.get("/api/relations")
def get_relations() -> Any:
    try:
        equipment = load_equipment_from_excel()
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except ValueError as e:
        return _error_response(str(e), 500)
    try:
        payload = _get_or_build_pipeline_sync(equipment)
        return jsonify(payload.get("relations", {}))
    except RuntimeError as e:
        return _error_response(str(e), 500)


@app.get("/api/files")
def get_files() -> Any:
    classified = classify_files(data_dir_path())
    out: Dict[str, Any] = {k: _path_like_to_json(v) for k, v in classified.items()}
    return jsonify(out)


@app.get("/api/layout/inspect")
def get_layout_inspect() -> Any:
    """Upload File Inspector: health of committed runtime plan.png (no pipeline run)."""
    info = inspect_runtime_layout(runtime_plan_path())
    return jsonify({"success": True, "layout_inspector": info})


@app.get("/api/status")
def get_status() -> Any:
    classified = classify_files(data_dir_path())
    required = {"layout", "excel"}
    missing = sorted(list(required - set(k for k, v in classified.items() if v)))
    files_out: Dict[str, Any] = {k: _path_like_to_json(v) for k, v in classified.items()}
    return jsonify(
        {
            "ready": len(missing) == 0,
            "missing": missing,
            "files": files_out,
        }
    )


@app.get("/api/pipeline")
def get_pipeline() -> Any:
    try:
        equipment = load_equipment_from_excel()
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except ValueError as e:
        return _error_response(str(e), 500)
    except Exception as e:  # noqa: BLE001
        return _error_response_struct("INVALID_EXCEL", _classify_pipeline_error(str(e)), TASK_VALIDATING, 400)
    try:
        return jsonify(_get_or_build_pipeline_sync(equipment))
    except AssetContractViolation as e:
        return _asset_violation_response(e, status_code=400)
    except RuntimeError as e:
        return _error_response(str(e), 500)
    except FileNotFoundError as e:
        return _error_response(str(e), 500)


@app.get("/api/layout_graph")
def get_layout_graph() -> Any:
    try:
        equipment = load_equipment_from_excel()
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except ValueError as e:
        return _error_response(str(e), 500)
    try:
        payload = _get_or_build_pipeline_sync(equipment)
        return jsonify(payload.get("layout_graph", {"nodes": [], "edges": [], "zones": [], "constraints": []}))
    except RuntimeError as e:
        return _error_response(str(e), 500)
    except FileNotFoundError as e:
        return _error_response(str(e), 500)
    except ValueError as e:
        return _error_response(str(e), 500)


@app.get("/api/pid_links")
def get_pid_links() -> Any:
    try:
        payload = _get_or_build_pipeline_sync()
        return jsonify(payload.get("phase_c", {}).get("pid_links", {}))
    except (RuntimeError, FileNotFoundError, ValueError) as e:
        return _error_response(str(e), 500)


@app.get("/api/topology")
def get_topology() -> Any:
    try:
        payload = _get_or_build_pipeline_sync()
        return jsonify(payload.get("phase_c", {}).get("topology_optimization", {}))
    except (RuntimeError, FileNotFoundError, ValueError) as e:
        return _error_response(str(e), 500)


@app.get("/health")
def get_health() -> Any:
    return jsonify({"status": "ok"})


@app.get("/api/plants")
def get_plants() -> Any:
    return jsonify(list_plants(runtime_dir_path()))


@app.get("/api/observability")
def get_observability_data() -> Any:
    rt = _RUNTIME_STATE.observability()
    return jsonify(
        {
            "active_tasks": rt.get("active_tasks", 0),
            "completed_tasks": rt.get("completed_tasks", 0),
            "failed_tasks": rt.get("failed_tasks", 0),
            "cache_hit_rate": rt.get("cache_hit_rate", 0.0),
            "avg_task_duration": rt.get("avg_task_duration", 0.0),
            "worker_status": rt.get("worker_status", "healthy"),
        }
    )


@app.get("/api/walls")
def get_walls() -> Any:
    try:
        payload = _get_or_build_pipeline_sync()
        walls_doc = payload.get("walls", {"walls": [], "rooms": [], "center": [0.0, 0.0]})
        return jsonify(walls_doc)
    except FileNotFoundError as e:
        return _error_response(str(e), 404)
    except RuntimeError as e:
        return _error_response(str(e), 500)


@app.post("/api/upload")
def upload_project_files() -> Any:
    runtime_dir = runtime_dir_path()
    runtime_dir.mkdir(parents=True, exist_ok=True)

    # Support explicit typed fields and arbitrary multi-file uploads.
    content_length = request.content_length or 0
    if content_length > MAX_UPLOAD_BYTES:
        return _error_response(f"Upload too large (>{MAX_UPLOAD_BYTES // (1024 * 1024)}MB)", 413)

    files = list(request.files.items(multi=True))
    if not files:
        return _error_response("No files uploaded", 400)

    plan_saved = False
    excel_saved = False
    plan_uploaded = False
    for field, storage in files:
        name = (storage.filename or "").lower()
        if not name:
            continue

        if field == "plan_file" or name.endswith((".png", ".jpg", ".jpeg", ".pdf")):
            target = runtime_dir / f"uploaded_layout{Path(name).suffix}"
            storage.save(target)
            plan_uploaded = True
            if target.suffix.lower() == ".pdf":
                try:
                    first_page_to_layout_png(target, runtime_plan_path())
                    ok_pdf, reason_pdf, _wh_pdf = validate_layout_raster(runtime_plan_path())
                    if not ok_pdf:
                        rp = runtime_plan_path()
                        try:
                            rp.unlink(missing_ok=True)  # type: ignore[call-arg]
                        except TypeError:
                            if rp.is_file():
                                rp.unlink()
                        _UPLOAD_LOG.warning(
                            "[SPATIAL_TRACE] layout pdf render invalid reason=%s file=%s",
                            reason_pdf,
                            name,
                        )
                        return _error_response_struct(
                            "INVALID_LAYOUT_IMAGE",
                            _layout_upload_validation_message(reason_pdf),
                            TASK_VALIDATING,
                            400,
                        )
                    plan_saved = True
                except Exception as e:
                    return _error_response(f"Invalid layout PDF: {e}", 400)
            else:
                ok, reason, _wh = validate_layout_raster(target)
                if not ok:
                    try:
                        target.unlink(missing_ok=True)  # type: ignore[call-arg]
                    except TypeError:
                        if target.is_file():
                            target.unlink()
                    _UPLOAD_LOG.warning(
                        "[SPATIAL_TRACE] layout rejected before commit reason=%s file=%s",
                        reason,
                        name,
                    )
                    return _error_response_struct(
                        "INVALID_LAYOUT_IMAGE",
                        _layout_upload_validation_message(reason),
                        TASK_VALIDATING,
                        400,
                    )
                runtime_plan_path().write_bytes(target.read_bytes())
                plan_rt = runtime_plan_path()
                ok2, reason2, _wh2 = validate_layout_raster(plan_rt)
                if not ok2:
                    try:
                        plan_rt.unlink(missing_ok=True)  # type: ignore[call-arg]
                    except TypeError:
                        if plan_rt.is_file():
                            plan_rt.unlink()
                    _UPLOAD_LOG.warning(
                        "[SPATIAL_TRACE] runtime plan invalid after copy reason=%s file=%s",
                        reason2,
                        name,
                    )
                    return _error_response_struct(
                        "INVALID_LAYOUT_IMAGE",
                        _layout_upload_validation_message(reason2),
                        TASK_VALIDATING,
                        400,
                    )
                plan_saved = True
            continue

        if field == "excel_file" or name.endswith(".xlsx"):
            target = runtime_excel_path()
            storage.save(target)
            excel_saved = True
            continue

        if field == "reference_file":
            target = runtime_dir / Path(storage.filename).name
            storage.save(target)
            continue

        if field == "structure_file":
            target = runtime_dir / f"uploaded_structure{Path(name).suffix}"
            storage.save(target)
            continue

        if field == "gad_file":
            target = runtime_dir / Path(storage.filename).name
            storage.save(target)
            continue

    if not plan_saved and runtime_plan_path().is_file():
        plan_saved = True
    if not excel_saved and runtime_excel_path().is_file():
        excel_saved = True
    if not (plan_saved and excel_saved):
        return _error_response("Missing required layout and/or excel file after upload", 400)

    cache_info = _invalidate_spatial_runtime_cache(clear_positions=plan_uploaded)

    task_id = _RUNTIME_STATE.new_task("Upload received")
    signature = _pipeline_signature()
    ctx = _register_task_context(task_id, signature)
    _RUNTIME_STATE.submit_pipeline_task(
        task_id,
        signature=signature,
        builder=lambda: _build_pipeline_dag(ctx),
    )
    return jsonify(
        {
            "success": True,
            "task_id": task_id,
            "status": TASK_QUEUED,
            "message": "Processing started",
            "cache": cache_info,
            "layout_inspector": inspect_runtime_layout(runtime_plan_path()),
        }
    )


@app.get("/api/task/<task_id>")
def get_task_status(task_id: str) -> Any:
    rec = _RUNTIME_STATE.get_task(task_id)
    if rec is None:
        return _error_response("Task not found", 404)
    if rec.get("status") == TASK_FAILED:
        raw_message = str(rec.get("error") or "")
        existing_code = str(rec.get("error_code") or "")
        stored_stage = str(rec.get("stage") or "")
        # Asset contract violation codes take priority and carry asset context.
        if existing_code in {"ASSET_MISSING", "ASSET_CORRUPTED"}:
            from backend.asset_contract import PLAN_IMAGE_CONTRACT

            error_payload: Dict[str, Any] = {
                "code": existing_code,
                "asset": PLAN_IMAGE_CONTRACT.name,
                "produced_by": PLAN_IMAGE_CONTRACT.produced_by,
                "consumed_by": "scene_render",
                "stage": "scene_render",
                "message": raw_message or (
                    f"Asset '{PLAN_IMAGE_CONTRACT.name}' contract violated."
                ),
            }
            return jsonify(
                {
                    "success": False,
                    "task_id": rec.get("task_id"),
                    "status": TASK_FAILED,
                    "progress": rec.get("progress"),
                    "message": rec.get("message"),
                    "error": error_payload,
                    "created_at": rec.get("created_at"),
                    "updated_at": rec.get("updated_at"),
                }
            )
        inferred_code, inferred_stage = _classify_pipeline_error_code(raw_message)
        code = inferred_code if inferred_code != "PIPELINE_ERROR" else (existing_code or inferred_code)
        stage = inferred_stage if inferred_code != "PIPELINE_ERROR" else (stored_stage or inferred_stage)
        return jsonify(
            {
                "success": False,
                "task_id": rec.get("task_id"),
                "status": TASK_FAILED,
                "progress": rec.get("progress"),
                "message": rec.get("message"),
                "error": _pipeline_error(
                    code or "PIPELINE_ERROR",
                    _classify_pipeline_error(raw_message),
                    stage or TASK_FAILED,
                ),
                "created_at": rec.get("created_at"),
                "updated_at": rec.get("updated_at"),
            }
        )
    return jsonify({"success": True, **rec})


@app.get("/api/task/latest")
def get_latest_task_status() -> Any:
    rec = _RUNTIME_STATE.latest_task()
    if rec is None:
        return jsonify({"success": True, "status": "idle", "progress": 0, "message": "No task yet"})
    return jsonify({"success": True, **rec})


@app.post("/api/task/<task_id>/cancel")
def cancel_task(task_id: str) -> Any:
    ctx = _get_task_context(task_id)
    rec = _RUNTIME_STATE.get_task(task_id)
    if rec is None:
        return _error_response_struct("TASK_NOT_FOUND", "Task not found", "cancelled", 404)
    if rec.get("status") in {TASK_DONE, TASK_FAILED, TASK_CANCELLED}:
        return jsonify({"success": True, "task_id": task_id, "status": rec.get("status"), "message": "Task already terminal"})
    if ctx and ctx.cancelled:
        ctx.cancelled.set()
    _transition(task_id, TASK_CANCELLED, int(rec.get("progress", 0)), "Cancelled by user")
    return jsonify({"success": True, "task_id": task_id, "status": TASK_CANCELLED})


@app.get("/api/upload/schema")
def get_upload_schema() -> Any:
    return jsonify(
        {
            "required": [
                {
                    "field": "plan_file",
                    "label": "Layout Drawing",
                    "accept": [".pdf", ".png", ".jpg", ".jpeg"],
                    "used_by": ["pdf_loader", "locator", "walls", "scene_engine"],
                },
                {
                    "field": "excel_file",
                    "label": "Equipment Excel",
                    "accept": [".xlsx"],
                    "used_by": ["excel_parser", "layout_graph", "scene_engine"],
                },
            ],
            "optional": [
                {
                    "field": "reference_file",
                    "label": "Reference Docs",
                    "accept": [".pdf"],
                    "used_by": ["pid_integration"],
                },
                {
                    "field": "structure_file",
                    "label": "Structure Drawing",
                    "accept": [".pdf", ".png", ".jpg", ".jpeg"],
                    "used_by": ["file_classifier", "walls (future weighting)"],
                },
            ],
            "developer_only": [
                {
                    "field": "gad_file",
                    "label": "GAD Typical",
                    "accept": [".pdf"],
                    "used_by": ["pid_integration (optional)"],
                }
            ],
        }
    )


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=False,
        use_reloader=False,
        threaded=True,
    )
